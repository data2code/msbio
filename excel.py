#!/usr/bin/env python
from __future__ import absolute_import
from __future__ import print_function
import sys
import pandas as pd
import openpyxl as oxl
import xlrd
import xlwt
import util
import os
import re
from six.moves import range

class Excel:
    """This class contains static methods to manipulate Excel files in either 2003 (.xls) or 2007 (.xlsx) format"""

    @staticmethod
    def read_xlsx(filename, sheet_name=None, sheet_index=None, l_allow_int_header=False):
        """Read an xlsx file, each sheet becomes a table.
        input and output are the same as read()"""
        file_format = os.path.splitext(filename)[-1]
        if file_format not in oxl.reader.excel.SUPPORTED_FORMATS:
            util.error_msg('File extension [%s] is not supported!\nModule openpyxl relies on file extension. If this is a valid file, please make sure file name has valid extension!' % file_format)
        wb=oxl.load_workbook(filename=filename) #, use_iterators=True)
        tables=[]
        headers=[]
        opts={}
        names=wb.get_sheet_names()
        opts['ActiveSheetIdx']=wb._active_sheet_index
        for i,s in enumerate(names):
            if sheet_name is not None and s!=sheet_name: continue
            if sheet_index is not None and i!=sheet_index: continue
            #print "*** "+s+" ***"
            ws=wb.get_sheet_by_name(s)
            data=[]
            nrow, ncol=(ws.max_row, ws.max_column)
            for row in range(1,nrow+1):
                data.append([ ws.cell(row=row, column=col).value for col in range(1,ncol+1)])
            #for row in ws.iter_rows(ws.calculate_dimension()):
            #    data.append([ cell.internal_value for cell in row])
            if len(data):
                if Excel.is_header(data[0], l_allow_int=l_allow_int_header):
                    header=data.pop(0)
                    headers.append(True)
                else:
                    header=["Col"+str(i+1) for i in range(len(data[0])) ]
                    headers.append(False)
                tables.append(pd.DataFrame(data=data, columns=header))
            else:
                tables.append(None)
        if len(names) and 'ActiveSheetIdx' not in opts:
            opts['ActiveSheetIdx']=0
        return (tables, names, headers, opts)

    @staticmethod
    def xls_proc_text(cell, value_proc=None, text_proc=None):
        """Converts the given cell to appropriate text."""
        """The proc will come in only when the given is value or text."""
        ttype = cell.ctype
        if ttype == xlrd.XL_CELL_EMPTY or ttype == xlrd.XL_CELL_TEXT or ttype == xlrd.XL_CELL_BLANK:
            if text_proc is None:
                return cell.value
            else:
                return text_proc(cell.value)
        if ttype == xlrd.XL_CELL_NUMBER or ttype == xlrd.XL_CELL_DATE or ttype == xlrd.XL_CELL_BOOLEAN:
            if value_proc is None:
                if cell.ctype in (2,3) and int(cell.value) == cell.value:
                    return int(cell.value)
                return cell.value
            else:
                return str(value_proc(cell.value))
        if cell.ctype == xlrd.XL_CELL_ERROR:
            # Apply no proc on this.
            return xlrd.error_text_from_code[cell.value]

    @staticmethod
    def read_xls(filename, sheet_name=None, sheet_index=None, l_allow_int_header=False):
        """Read an xls file, each sheet becomes a table.
        input and output are the same as read()"""
        wb=xlrd.open_workbook(filename=filename)
        tables=[]
        headers=[]
        names=wb.sheet_names()
        opts={}
        for i,s in enumerate(wb.sheets()):
            #print i, s.name, s.sheet_selected, s.sheet_visible
            if s.sheet_selected:
                opts['ActiveSheetIdx']=i
                break
        for i,s in enumerate(names):
            if sheet_name is not None and s!=sheet_name: continue
            if sheet_index is not None and i!=sheet_index: continue
            #print "*** "+s+" ***"
            ws=wb.sheet_by_index(i)
            data=[]
            for rx in range(ws.nrows):
                data.append([ Excel.xls_proc_text(ws.cell(rx, cx)) for cx in range(ws.ncols) ])
            if len(data):
                if Excel.is_header(data[0], l_allow_int=l_allow_int_header):
                    header=data.pop(0)
                    headers.append(True)
                else:
                    header=["Col"+str(i+1) for i in range(len(data[0])) ]
                    headers.append(False)
                tables.append(pd.DataFrame(data=data, columns=header))
            else:
                tables.append(None)
        if len(names) and 'ActiveSheetIdx' not in opts:
            opts['ActiveSheetIdx']=0
        #print tables
        return (tables, names, headers, opts)

    @staticmethod
    def write_xlsx(filename, tables, names=None, headers=None):
        """Similar to write(), but writes to xlsx format"""
        if names is None:
            names=["Sheet"+str(i+1) for i in range(len(tables)) ]
        if headers is None:
            headers=[ True for i in range(len(tables)) ]
        #wb=oxl.Workbook(optimized_write=True)
        wb=oxl.Workbook(write_only=True)
        for i,t in enumerate(tables):
            ws=wb.create_sheet(index=i, title=names[i])
            if t is None: continue
            if headers[i]:
                ws.append(t.header())
            for j in range(len(t)):
                data=[None if pd.isnull(x) else x for x in list(t.iloc[j])]
                # API does not like numeric np.nan or np.inf ..., Excel format will be considered invalid
                ws.append(data)
        wb.save(filename)

    @staticmethod
    def write_xls(filename, tables, names=None, headers=None):
        """Similar to write(), but writes to xls format"""
        if names is None:
            names=["Sheet"+str(i+1) for i in range(len(tables)) ]
        if headers is None:
            headers=[ True for i in range(len(tables)) ]
        wb=xlwt.Workbook()
        for i,t in enumerate(tables):
            ws=wb.add_sheet(names[i])
            if t is None: continue
            offset=0
            S=t.header()
            ncols=len(S)
            if headers[i]:
                for cx in range(ncols):
                    ws.write(0, cx, S[cx])
                offset=1
            for rx in range(len(t)):
                for cx in range(ncols):
                    ws.write(rx+offset, cx, t.iat[rx, cx])
        wb.save(filename)

    @staticmethod
    def is_header(row, l_allow_int=False):
        """Check if a row can be column header
        row: list, containing the cells of the first row
        l_allow_int: boolean, default False, whether int can be used as a column name
        return boolean"""
        if len(row) != len(util.unique(row)):
            return False
        for x in row:
            if type(x) is float or type(x) is complex: return False
            if (type(x) is int) and not l_allow_int: return False
            if not x: return False
        return True

    @staticmethod
    def xls2xlsx(infile, outfile=None):
        """Converts an xls file into an xlsx file. If outfile is not provides, use the same file name, but .xlsx as the extension"""
        if outfile is None:
            outfile=re.sub(r'\.xls', '.xlsx', infile, re.IGNORECASE)
        tables, names, headers = Excel.read_xls(infile)
        Excel.write_xlsx(outfile, tables, names, headers)

    @staticmethod
    def xlsx2xls(infile, outfile=None):
        """Converts an xlsx file into an xls file. If outfile is not provides, use the same file name, but .xls as the extension"""
        if outfile is None:
            outfile=re.sub(r'\.xlsx', '.xls', infile, re.IGNORECASE)
        tables, names, headers = Excel.read_xlsx(infile)
        Excel.write_xls(outfile, tables, names, headers)

    @staticmethod
    def is_xlsx(filename):
        """Check if a file is XLSX format
        filename: str. If reads the first two bytes of the file. XLSX is zip format, where first two bytes are 'PK'.
        return boolean"""
        s=""
        f=open(filename, "rb")
        try:
            s=f.read(2)
        finally:
            f.close()
        if type(s)!=str: s=s.decode('ascii', 'ignore')
        return s=="PK"
        if sys.version_info[0]>=3: # python3
            return s==b"PK"
        else:
            return s=="PK"

    @staticmethod
    def read(filename, sheet_name=None, sheet_index=None, l_allow_int_header=False):
        """Read an xlsx/xls file, each sheet becomes a table.
        filename: str, filename
        sheet_name: str, default None. If specified, it only reads the sheet that matches the name
        sheet_index: int, default None. If specified, it only reads the particular sheet. First sheet has index 0.
        l_allow_int_header: boolean, default False. For each sheet, it converts the data into a DataFrame object
            The column header of the table may be taken from the first row, if the row looks like a header row
            A row containing float number is not a header. This option specifies whether int can be used as a column header
            If no column header, table columns will be Col1, Col2, etc.
        return tuple: (tables, names, headers). tables is a list of tables. Empty sheet gives a None table object.
            names is the list of sheet names. headers is a list of boolean, indicating whether column header comes from sheet"""
        if Excel.is_xlsx(filename):
            return Excel.read_xlsx(filename, sheet_name, sheet_index, l_allow_int_header)
        else:
            return Excel.read_xls(filename, sheet_name, sheet_index, l_allow_int_header)

    @staticmethod
    def write(filename, tables, names=None, headers=None, format='XLSX'):
        """Write an xlsx/xls file, each table becomes a sheet.
        filename: str, filename
        tables: dataframe. If a sheet is empty, put None.
        names: list[str], default None. If specified, it specifies sheet names
        headers: list[boolean], default None. If specified, it specifies whether the output table column names
        format: 'XLSX' or 'XLS'
        return none: writes the file"""
        if format=='XLSX':
            Excel.write_xlsx(filename, tables, names, headers)
        else:
            Excel.write_xls(filename, tables, names, headers)

if __name__=="__main__":
    fn = r"~\Test file 3.xls"
    tables, names, headers, opts = Excel.read(fn)


    exit()
    tables, names, headers, opts=Excel.read_xls('2.xlsx')
    for i,x in enumerate(tables):
        t = x.fillna(value='')
        print(">>> ", names[i], headers[i])
        print(t)
    exit()
    Excel.read('NorthWind.xlsx')
    print(Excel.is_xlsx('NorthWind.xlsx'))
    print(Excel.is_xlsx('NorthWind.xls'))
    exit()
    Excel.xlsx2xls('t.xlsx')
    Excel.xls2xlsx('t.xls')
    exit()
    tables, names, headers=Excel.read_xlsx('NorthWind.xlsx')
    for i,x in enumerate(tables):
        print(">>> ", names[i], headers[i])
        print(x[:5])

    tables, names, headers=Excel.read_xls('NorthWind.xls')
    for i,x in enumerate(tables):
        print(">>> ", names[i], headers[i])
        print(x[:5])

    Excel.write_xlsx('t.xlsx', tables, names, headers)
    Excel.write_xls('t.xls', tables, names, headers)
